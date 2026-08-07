from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

wb = Workbook()

# ─── COLOR PALETTE ───────────────────────────────────────────────────────────
C_NAVY   = "1B3A5C"
C_TEAL   = "006D77"
C_GREEN  = "1A7A41"
C_RED    = "C0392B"
C_ORANGE = "D35400"
C_PURPLE = "5B3FA0"
C_WHITE  = "FFFFFF"
C_LGRAY  = "F2F4F6"
C_MGRAY  = "D5D8DC"
C_DKGRAY = "717D7E"
C_PASS_BG = "D5F5E3"
C_PASS_FG = "145A32"
C_FAIL_BG = "FADBD8"
C_FAIL_FG = "922B21"
C_EXCL_BG = "EDE7F6"
C_EXCL_FG = "4527A0"
C_WARN_BG = "FEF9E7"
C_HPV16   = "D6EAF8"
C_HPV18   = "D5F5E3"
C_HPV45   = "E8DAEF"
C_HPVOTH  = "FEF9E7"
C_NEG_BG  = "F0F3F4"


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=10, italic=False, name="Arial"):
    return Font(bold=bold, color=color or "000000", size=size,
                italic=italic, name=name)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def thin_border():
    return border("thin")

def apply_header(ws, row, col, value, bg=C_NAVY, fg=C_WHITE, bold=True,
                  h="center", fontsize=10, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=fontsize)
    cell.alignment = align(h, wrap=wrap)
    cell.border = thin_border()
    return cell

def apply_cell(ws, row, col, value, bg=None, fg="000000", bold=False,
               h="center", border_on=True, wrap=False, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = align(h, wrap=wrap)
    if border_on:
        cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell

def merge_header(ws, row, c1, c2, value, bg=C_NAVY, fg=C_WHITE, fontsize=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=fontsize)
    cell.alignment = align("center")
    return cell

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: สรุป (Summary)
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "สรุป"
ws1.sheet_view.showGridLines = False

# Title block
ws1.merge_cells("A1:L1")
c = ws1["A1"]
c.value = "รายงาน EQA HPV DNA — ห้องปฏิบัติการ HB16305 | ร.พ.พระจอมเกล้า เพชรบุรี"
c.fill = fill(C_NAVY)
c.font = font(bold=True, color=C_WHITE, size=14)
c.alignment = align("center")
row_height(ws1, 1, 28)

ws1.merge_cells("A2:L2")
c = ws1["A2"]
c.value = "cobas® HPV (Roche Cobas 5800/6800/8800)  |  QCMD: QAV094130 · TH193  |  สวส.กรมวิทยาศาสตร์การแพทย์: HB16305  |  ปี 2567–2569"
c.fill = fill("154360")
c.font = font(color=C_WHITE, size=9)
c.alignment = align("center")
row_height(ws1, 2, 16)

# ── Section A: QCMD Summary ──────────────────────────────────────────────────
merge_header(ws1, 4, 1, 6, "QCMD — QAV094130 (TH193)", bg=C_TEAL)
merge_header(ws1, 4, 7, 12, "สวส. กรมวิทยาศาสตร์การแพทย์ — HB16305", bg=C_PURPLE)
row_height(ws1, 4, 20)

# Sub-headers
for col, hdr in enumerate(["รายการ", "ค่า", "", "หมายเหตุ", "", ""], start=1):
    if hdr:
        apply_header(ws1, 5, col, hdr, bg="1A607A")
for col, hdr in enumerate(["รายการ", "ค่า", "", "หมายเหตุ", "", ""], start=7):
    if hdr:
        apply_header(ws1, 5, col, hdr, bg="3D2680")

qcmd_rows = [
    ("จำนวน Challenges", "5 รอบ (2024–2026)", "", "QCMD HPV PreservCyt"),
    ("ผลการประเมิน", "5/5 PASS", "", "Core Panel Score: 0 — Highly Satisfactory"),
    ("Sensitivity", "100%", "", "24/24 TP samples"),
    ("Specificity", "100%", "", "5/5 TN samples"),
    ("Average peer %", "98.4%", "", "ช่วง: 98.2%–99.1%"),
]
svs_rows = [
    ("จำนวนรอบประเมิน", "5 รอบ (2567–2569)", "", "สวส. HPV DNA EQA"),
    ("ผลการประเมิน", "4/5 PASS — 1 FAIL", "", "2568 C1: ไม่ผ่านเกณฑ์"),
    ("Sensitivity", "93.8%", "", "15/16 TP (FN: 2568C1 HPV-01)"),
    ("Specificity", "100%", "", "7/7 TN samples"),
    ("คะแนนรวม", "46/50 (92%)", "", "เกณฑ์ยอมรับ: 100% ต่อรอบ"),
]

for i, (lbl, val, _, note) in enumerate(qcmd_rows):
    r = 6 + i
    is_pass = "PASS" in val and "FAIL" not in val
    v_col = "1A7A41" if is_pass else ("C0392B" if "FAIL" in val else "000000")
    apply_cell(ws1, r, 1, lbl, h="left")
    apply_cell(ws1, r, 2, val, bold=True, fg=v_col)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    apply_cell(ws1, r, 3, note, fg=C_DKGRAY, h="left", wrap=True)
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    apply_cell(ws1, r, 5, "")

for i, (lbl, val, _, note) in enumerate(svs_rows):
    r = 6 + i
    is_pass = val == "4/5 PASS — 1 FAIL" or "PASS" in val
    v_col = "C0392B" if "FAIL" in val else ("1A7A41" if "100%" in val or "PASS" in val else "000000")
    apply_cell(ws1, r, 7, lbl, h="left")
    apply_cell(ws1, r, 8, val, bold=True, fg=v_col)
    ws1.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
    apply_cell(ws1, r, 9, note, fg=C_DKGRAY, h="left", wrap=True)
    ws1.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
    apply_cell(ws1, r, 11, "")

for r in range(6, 11):
    row_height(ws1, r, 18)

# ── FAIL Alert ───────────────────────────────────────────────────────────────
ws1.merge_cells("A12:L12")
c = ws1["A12"]
c.value = "⚠  สวส. 2568 ครั้งที่ 1 — ไม่ผ่านเกณฑ์ (FAIL) | HPV-01 (HPV 66 · Non-16,18): ห้องปฏิบัติการรายงาน Negative — False Negative | คะแนน 8/10 (80%) — เกณฑ์ 100%"
c.fill = fill(C_FAIL_BG)
c.font = font(bold=True, color=C_FAIL_FG, size=10)
c.alignment = align("left", wrap=True)
row_height(ws1, 12, 22)

# ── Combined Comparison Table ─────────────────────────────────────────────────
row_height(ws1, 13, 8)
merge_header(ws1, 14, 1, 12, "ตารางเปรียบเทียบ EQA — ทุก challenge/รอบ", bg=C_NAVY)
row_height(ws1, 14, 20)

comp_hdrs = ["Program","ปี","รอบ","รหัสรายงาน","คะแนน/Score","ผลประเมิน",
             "Sensitivity","Specificity","TP","TN","FN","หมายเหตุ"]
for c2, h in enumerate(comp_hdrs, start=1):
    apply_header(ws1, 15, c2, h, bg="234F6E")
row_height(ws1, 15, 16)

comp_data = [
    ("QCMD","2567","C1","QAV094130-2024C1","Score: 0","PASS","100%","100%",4,1,0,"Highly Satisfactory"),
    ("QCMD","2567","C2","QAV094130-2024C2","Score: 0","PASS","100%","100%",5,1,0,"Highly Satisfactory"),
    ("สวส.","2567","C1","67-021-146","8/8 (100%)*","PASS","100%","100%",2,2,0,"HPV05 ยกเว้น (ค่าพ้องไม่ผ่าน)"),
    ("สวส.","2567","C2","67-022-146","10/10 (100%)","PASS","100%","100%",3,2,0,""),
    ("QCMD","2568","C1","QAV094130-2025C1","Score: 0","PASS","100%","100%",5,1,0,"Highly Satisfactory"),
    ("QCMD","2568","C2","QAV094130-2025C2","Score: 0","PASS","100%","100%",5,1,0,"Highly Satisfactory"),
    ("สวส.","2568","C1","68-021-063","8/10 (80%)","FAIL","75%","100%",3,1,1,"HPV-01 HPV66→Negative (FN) ไม่ผ่านเกณฑ์"),
    ("สวส.","2568","C2","68-022-164","10/10 (100%)*","PASS","100%","100%",3,1,0,"HPV04 ยกเว้น (ค่าพ้องไม่ผ่าน)"),
    ("QCMD","2569","C1","QAV094130-2026C1","Score: 0","PASS","100%","100%",5,1,0,"Highly Satisfactory"),
    ("สวส.","2569","C1","69-021-120","10/10 (100%)","PASS","100%","100%",4,1,0,""),
]

for i, row_data in enumerate(comp_data):
    r = 16 + i
    is_fail = row_data[5] == "FAIL"
    is_svs = row_data[0] == "สวส."
    row_bg = C_FAIL_BG if is_fail else (None)
    prog_bg = C_EXCL_BG if is_svs else "E0F7F4"
    prog_fg = C_PURPLE if is_svs else C_TEAL

    apply_cell(ws1, r, 1, row_data[0], bg=prog_bg, fg=prog_fg, bold=True)
    for c2, val in enumerate(row_data[1:], start=2):
        col_bg = row_bg
        col_fg = "000000"
        col_bold = False
        if c2 == 6:  # ผล
            col_bg = C_PASS_BG if val == "PASS" else C_FAIL_BG
            col_fg = C_PASS_FG if val == "PASS" else C_FAIL_FG
            col_bold = True
        elif c2 == 7:  # Sensitivity
            col_fg = C_PASS_FG if val == "100%" else C_FAIL_FG
            col_bold = True
        apply_cell(ws1, r, c2, val, bg=col_bg, fg=col_fg, bold=col_bold, h="center")

    row_height(ws1, r, 16)

apply_cell(ws1, 26, 1, "* ไม่นับตัวอย่างที่ค่าพ้องกลุ่มไม่เป็นไปตามเกณฑ์",
           h="left", fg=C_DKGRAY, border_on=False)

set_col_widths(ws1, {1:14,2:6,3:6,4:20,5:18,6:12,7:12,8:12,9:5,10:5,11:5,12:38})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: QCMD
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("QCMD")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:J1")
c = ws2["A1"]
c.value = "QCMD HPV PreservCyt EQA — QAV094130 (TH193) | HB16305 | cobas® HPV | 2024–2026"
c.fill = fill(C_TEAL)
c.font = font(bold=True, color=C_WHITE, size=13)
c.alignment = align("center")
row_height(ws2, 1, 26)

# KPI row
kpis = [("Challenges","5"),("PASS","5/5"),("Sensitivity","100%"),("Specificity","100%"),("Avg peer %","98.4%")]
for i, (lbl, val) in enumerate(kpis):
    c1 = 1 + i*2
    ws2.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c1+1)
    cell = ws2.cell(row=2, column=c1, value=f"{lbl}: {val}")
    cell.fill = fill("154360")
    cell.font = font(bold=True, color=C_WHITE, size=10)
    cell.alignment = align("center")
row_height(ws2, 2, 18)

# Challenge data
challenges = [
    {
        "label": "2024 C1 (พ.ค. 2567)", "date": "10/05/2567",
        "datasets": 191, "countries": 30, "peer_pct": "99.1%", "core_score": 0,
        "result": "PASS",
        "samples": [
            ("24C1-01","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","29.70","99.0%"),
            ("24C1-02","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","26.44","99.5%"),
            ("24C1-03","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","26.53","99.0%"),
            ("24C1-04","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","23.51","99.0%"),
            ("24C1-05","HPV16 Caski (low)","HPV16","EDU","HPV 16","Positive","0","31.05","84.8%"),
            ("24C1-06","Negative","Neg","CORE","Negative","Negative","0","—","99.0%"),
        ]
    },
    {
        "label": "2024 C2 (ต.ค. 2567)", "date": "04/10/2567",
        "datasets": 196, "countries": 30, "peer_pct": "98.5%", "core_score": 0,
        "result": "PASS",
        "samples": [
            ("24C2-01","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","25.49","98.5%"),
            ("24C2-02","HPV45 CC10b","HPV45","CORE","HPV 45","Positive","0","26.63","99.0%"),
            ("24C2-03","Negative","Neg","CORE","Negative","Negative","0","—","98.0%"),
            ("24C2-04","HPV45 CC10b","HPV45","CORE","HPV 45","Positive","0","26.54","98.5%"),
            ("24C2-05","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","26.75","98.0%"),
            ("24C2-06","HPV16+HPV18","HPV16+18","CORE","HPV 16+18","Positive","0","27.16","99.0%"),
        ]
    },
    {
        "label": "2025 C1 (พ.ค. 2568)", "date": "05/2568",
        "datasets": 256, "countries": 29, "peer_pct": "98.5%", "core_score": 0,
        "result": "PASS",
        "samples": [
            ("25C1-01","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","27.26","100.0%"),
            ("25C1-02","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","27.37","98.8%"),
            ("25C1-03","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","23.43","99.6%"),
            ("25C1-04","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","27.21","99.6%"),
            ("25C1-05","Negative","Neg","CORE","Negative","Negative","0","—","98.0%"),
            ("25C1-06","HPV18 low","HPV18","CORE","HPV 18","Positive","0","30.15","94.9%"),
        ]
    },
    {
        "label": "2025 C2 (ต.ค. 2568)", "date": "10/2568",
        "datasets": 253, "countries": 29, "peer_pct": "98.7%", "core_score": 0,
        "result": "PASS",
        "samples": [
            ("25C2-01","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","25.39","99.2%"),
            ("25C2-02","HPV45 CC10b","HPV45","CORE","HPV 45","Positive","0","27.70","98.4%"),
            ("25C2-03","HPV18 Hela","HPV18","CORE","HPV 18","Positive","0","26.48","98.0%"),
            ("25C2-04","Negative","Neg","CORE","Negative","Negative","0","—","98.8%"),
            ("25C2-05","HPV45 CC10b","HPV45","CORE","HPV 45","Positive","0","27.55","98.4%"),
            ("25C2-06","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","24.91","99.2%"),
        ]
    },
    {
        "label": "2026 C1 (พ.ค. 2569)", "date": "05/2569",
        "datasets": 225, "countries": 31, "peer_pct": "98.2%", "core_score": 0,
        "result": "PASS",
        "samples": [
            ("26C1-01","HPV16 Caski (low)","HPV16","CORE","HPV 16","Positive","0","28.41","96.9%"),
            ("26C1-02","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","25.63","99.1%"),
            ("26C1-03","HPV16 Caski","HPV16","CORE","HPV 16","Positive","0","26.16","99.1%"),
            ("26C1-04","HPV45 CC10b","HPV45","CORE","HPV 45","Positive","0","26.64","98.7%"),
            ("26C1-05","HPV16 Caski (low)","HPV16","CORE","HPV 16","Positive","0","29.99","97.3%"),
            ("26C1-06","Negative","Neg","CORE","Negative","Negative","0","—","98.2%"),
        ]
    },
]

TYPE_BG = {"HPV16": C_HPV16, "HPV18": C_HPV18, "HPV45": C_HPV45,
           "HPV16+18": C_HPVOTH, "Neg": C_NEG_BG}
SAMPLE_HDRS = ["Sample ID","Cell Line / Content","Type","Status",
               "Expected Result","Lab Result","Core Score","CT Value","Peer %"]

current_row = 4
for ch in challenges:
    # Challenge header
    ws2.merge_cells(start_row=current_row, start_column=1,
                    end_row=current_row, end_column=10)
    cell = ws2.cell(row=current_row, column=1,
                    value=f"Challenge: {ch['label']}   |   วันที่: {ch['date']}   |   ผล: {ch['result']}   |   Core Panel Score: {ch['core_score']}")
    cell.fill = fill(C_NAVY)
    cell.font = font(bold=True, color=C_WHITE, size=10)
    cell.alignment = align("left")
    row_height(ws2, current_row, 18)
    current_row += 1

    # Summary row
    summary_cols = [
        ("Datasets",str(ch["datasets"])),("Countries",str(ch["countries"])),
        ("Core Score",str(ch["core_score"])),("Avg Peer %",ch["peer_pct"]),
        ("Sensitivity","100%"),("Specificity","100%"),
    ]
    for ci, (lbl, val) in enumerate(summary_cols):
        col = 1 + ci*1
        if ci < 6:
            apply_header(ws2, current_row, ci+1, f"{lbl}: {val}",
                        bg="2A9D8F" if ci % 2 == 0 else C_TEAL, fontsize=9)
    # fill remaining cols
    for ci in range(7, 11):
        apply_cell(ws2, current_row, ci, "", bg="2A9D8F")
    row_height(ws2, current_row, 15)
    current_row += 1

    # Column headers
    for ci, h in enumerate(SAMPLE_HDRS, start=1):
        apply_header(ws2, current_row, ci, h, bg="234F6E", fontsize=9)
    row_height(ws2, current_row, 15)
    current_row += 1

    # Sample rows
    for s in ch["samples"]:
        sid, content, htype, status, expected, lab_result, score, ct, peer = s
        row_bg = None
        edu_row = status == "EDU"
        low_peer = False
        try:
            peer_f = float(peer.replace("%", ""))
            if peer_f < 95:
                low_peer = True
        except:
            pass

        type_bg = TYPE_BG.get(htype, "FFFFFF")
        apply_cell(ws2, current_row, 1, sid, bg=C_LGRAY, bold=True, h="center")
        apply_cell(ws2, current_row, 2, content, h="left")
        apply_cell(ws2, current_row, 3, htype, bg=type_bg, bold=True)
        apply_cell(ws2, current_row, 4, status,
                   bg=C_WARN_BG if edu_row else C_PASS_BG,
                   fg="9A7D0A" if edu_row else C_PASS_FG)
        apply_cell(ws2, current_row, 5, expected, h="left")
        apply_cell(ws2, current_row, 6, lab_result, bg=C_PASS_BG, fg=C_PASS_FG, bold=True)
        apply_cell(ws2, current_row, 7, int(score), bg=C_PASS_BG, fg=C_PASS_FG, bold=True)
        apply_cell(ws2, current_row, 8, ct, bg=C_WARN_BG if (ct != "—" and float(ct.replace("—","0")) > 30) else None)
        apply_cell(ws2, current_row, 9, peer,
                   bg=C_WARN_BG if low_peer else None,
                   fg=C_ORANGE if low_peer else C_PASS_FG)
        row_height(ws2, current_row, 15)
        current_row += 1

    current_row += 1  # spacer

set_col_widths(ws2, {1:13,2:22,3:10,4:8,5:20,6:20,7:10,8:10,9:10})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: สวส.
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("สวส.")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
c = ws3["A1"]
c.value = "สวส. กรมวิทยาศาสตร์การแพทย์ — HPV DNA EQA | HB16305 | cobas® HPV | ปี 2567–2569"
c.fill = fill(C_PURPLE)
c.font = font(bold=True, color=C_WHITE, size=13)
c.alignment = align("center")
row_height(ws3, 1, 26)

kpis3 = [("รอบทั้งหมด","5 รอบ"),("PASS","4/5"),("FAIL","1 รอบ"),
          ("Sensitivity","93.8%"),("Specificity","100%")]
for i, (lbl, val) in enumerate(kpis3):
    c1 = 1 + i*2
    ws3.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c1+1)
    cell = ws3.cell(row=2, column=c1, value=f"{lbl}: {val}")
    cell.fill = fill(C_FAIL_BG if val == "1 รอบ" else "3D2680")
    cell.font = font(bold=True,
                     color=C_FAIL_FG if val == "1 รอบ" else C_WHITE, size=10)
    cell.alignment = align("center")
row_height(ws3, 2, 18)

ws3.merge_cells("A3:J3")
c = ws3["A3"]
c.value = ("⚠  2568 ครั้งที่ 1 — FAIL: HPV-01 (HPV 66 · Non-16,18) รายงาน Negative → 0/2 คะแนน "
           "| รวม 8/10 (80%) | เกณฑ์ยอมรับ 100%")
c.fill = fill(C_FAIL_BG)
c.font = font(bold=True, color=C_FAIL_FG, size=10)
c.alignment = align("left", wrap=True)
row_height(ws3, 3, 22)

svs_rounds = [
    {
        "label": "2567 ครั้งที่ 1 (พ.ค. 2567)",
        "report": "67-021-146", "date": "10/05/2567",
        "score": "8/8 (100%)*", "result": "PASS",
        "note": "HPV-05 ยกเว้น (ค่าพ้องกลุ่มไม่ผ่านเกณฑ์)",
        "samples": [
            ("HPV-01","HPV 52 (Non-16,18)","ประเมิน","HPV Non-16,18","2/2","TP",""),
            ("HPV-02","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-03","HPV 16 + HPV 66","ประเมิน","HPV 16 + HPV Non-16,18","2/2","TP",""),
            ("HPV-04","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-05","HPV 58 + HPV 39","ยกเว้น","HPV Non-16,18","—","EXCL","ค่าพ้องกลุ่มไม่เป็นไปตามเกณฑ์"),
        ]
    },
    {
        "label": "2567 ครั้งที่ 2 (ก.ย. 2567)",
        "report": "67-022-146", "date": "04/09/2567",
        "score": "10/10 (100%)", "result": "PASS",
        "note": "",
        "samples": [
            ("HPV-01","HPV 16","ประเมิน","HPV 16","2/2","TP",""),
            ("HPV-02","HPV 52 (Non-16,18)","ประเมิน","HPV Non-16,18","2/2","TP",""),
            ("HPV-03","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-04","HPV 68 (Non-16,18)","ประเมิน","HPV Non-16,18","2/2","TP",""),
            ("HPV-05","Negative","ประเมิน","Negative","2/2","TN",""),
        ]
    },
    {
        "label": "2568 ครั้งที่ 1 (ก.พ. 2568)  ⚠ FAIL",
        "report": "68-021-063", "date": "24/02/2568",
        "score": "8/10 (80%)", "result": "FAIL",
        "note": "HPV-01: HPV 66 (Non-16,18) → Negative (False Negative) — ไม่ผ่านเกณฑ์",
        "samples": [
            ("HPV-01","HPV 66 (Non-16,18)","ประเมิน","Negative  ← FALSE NEGATIVE!","0/2","FN","ไม่รู้จัก HPV66 → รายงาน Negative"),
            ("HPV-02","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-03","HPV 16","ประเมิน","HPV 16","2/2","TP",""),
            ("HPV-04","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
            ("HPV-05","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
        ]
    },
    {
        "label": "2568 ครั้งที่ 2 (ก.ค. 2568)",
        "report": "68-022-164", "date": "22/07/2568",
        "score": "10/10 (100%)*", "result": "PASS",
        "note": "HPV-04 ยกเว้น (ค่าพ้องกลุ่มไม่ผ่านเกณฑ์)",
        "samples": [
            ("HPV-01","HPV 52 (Non-16,18)","ประเมิน","HPV Non-16,18","2/2","TP",""),
            ("HPV-02","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
            ("HPV-03","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-04","HPV 58 (Non-16,18)","ยกเว้น","Negative","—","EXCL","ค่าพ้องกลุ่มไม่เป็นไปตามเกณฑ์"),
            ("HPV-05","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
        ]
    },
    {
        "label": "2569 ครั้งที่ 1 (ก.พ. 2569)",
        "report": "69-021-120", "date": "20/02/2569",
        "score": "10/10 (100%)", "result": "PASS",
        "note": "",
        "samples": [
            ("HPV-01","HPV 16","ประเมิน","HPV 16","2/2","TP",""),
            ("HPV-02","Negative","ประเมิน","Negative","2/2","TN",""),
            ("HPV-03","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
            ("HPV-04","HPV 52 (Non-16,18)","ประเมิน","HPV Non-16,18","2/2","TP",""),
            ("HPV-05","HPV 18","ประเมิน","HPV 18","2/2","TP",""),
        ]
    },
]

SVS_HDRS = ["รหัสตัวอย่าง","ค่ากำหนด","สถานะ","ผลห้องปฏิบัติการ","คะแนน","TP/TN/FP/FN","หมายเหตุ"]

cur3 = 5
for rd in svs_rounds:
    is_fail_round = rd["result"] == "FAIL"
    hdr_bg = C_RED if is_fail_round else C_PURPLE

    ws3.merge_cells(start_row=cur3, start_column=1, end_row=cur3, end_column=10)
    cell = ws3.cell(row=cur3, column=1,
                    value=f"รอบ: {rd['label']}  |  รายงาน: {rd['report']}  |  วันที่: {rd['date']}  |  คะแนน: {rd['score']}  |  ผล: {rd['result']}")
    cell.fill = fill(hdr_bg)
    cell.font = font(bold=True, color=C_WHITE, size=10)
    cell.alignment = align("left")
    row_height(ws3, cur3, 18)
    cur3 += 1

    if rd["note"]:
        ws3.merge_cells(start_row=cur3, start_column=1, end_row=cur3, end_column=10)
        nc = ws3.cell(row=cur3, column=1, value=f"หมายเหตุ: {rd['note']}")
        nc.fill = fill(C_FAIL_BG if is_fail_round else C_WARN_BG)
        nc.font = font(color=C_FAIL_FG if is_fail_round else "7D6608", size=9)
        nc.alignment = align("left")
        row_height(ws3, cur3, 15)
        cur3 += 1

    for ci, h in enumerate(SVS_HDRS, start=1):
        apply_header(ws3, cur3, ci, h, bg="3D2680", fontsize=9)
    ws3.merge_cells(start_row=cur3, start_column=8, end_row=cur3, end_column=10)
    row_height(ws3, cur3, 15)
    cur3 += 1

    for s in rd["samples"]:
        sid, expected, status, lab_result, score, verdict, note = s
        is_excl = status == "ยกเว้น"
        is_fn = verdict == "FN"
        is_tp = verdict == "TP"
        is_tn = verdict == "TN"

        row_alpha = 0.5 if is_excl else 1.0
        row_bg2 = C_FAIL_BG if is_fn else (C_EXCL_BG if is_excl else None)

        apply_cell(ws3, cur3, 1, sid, bg=C_LGRAY, bold=True)
        apply_cell(ws3, cur3, 2, expected, h="left",
                   bg=C_FAIL_BG if is_fn else (C_EXCL_BG if is_excl else None))
        apply_cell(ws3, cur3, 3, status,
                   bg=C_EXCL_BG if is_excl else C_PASS_BG,
                   fg=C_EXCL_FG if is_excl else C_PASS_FG)
        apply_cell(ws3, cur3, 4, lab_result, h="left",
                   bg=C_FAIL_BG if is_fn else (C_EXCL_BG if is_excl else C_PASS_BG),
                   fg=C_FAIL_FG if is_fn else (C_EXCL_FG if is_excl else C_PASS_FG),
                   bold=is_fn)
        apply_cell(ws3, cur3, 5, score, bold=True,
                   fg=C_FAIL_FG if score == "0/2" else (C_DKGRAY if score == "—" else C_PASS_FG))
        verdict_bg = {"TP": C_PASS_BG, "TN": C_HPV16, "FN": C_FAIL_BG,
                      "EXCL": C_EXCL_BG}.get(verdict, None)
        verdict_fg = {"TP": C_PASS_FG, "TN": "154360", "FN": C_FAIL_FG,
                      "EXCL": C_EXCL_FG}.get(verdict, "000000")
        apply_cell(ws3, cur3, 6, verdict, bg=verdict_bg, fg=verdict_fg, bold=True)
        ws3.merge_cells(start_row=cur3, start_column=7, end_row=cur3, end_column=10)
        apply_cell(ws3, cur3, 7, note, h="left", fg=C_DKGRAY)
        row_height(ws3, cur3, 15)
        cur3 += 1

    cur3 += 1  # spacer

ws3.cell(row=cur3, column=1,
         value="* ไม่นับตัวอย่างที่ค่าพ้องกลุ่มไม่เป็นไปตามเกณฑ์  |  TP=True Positive · TN=True Negative · FN=False Negative · EXCL=Excluded"
         ).font = font(color=C_DKGRAY, size=9)

set_col_widths(ws3, {1:12,2:28,3:10,4:36,5:8,6:10,7:40})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Raw Data
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Raw Data")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A2"

ws4.merge_cells("A1:L1")
c = ws4["A1"]
c.value = "Raw Data — ทุก Sample ทั้งสองโปรแกรม | HB16305 | 2567–2569"
c.fill = fill(C_NAVY)
c.font = font(bold=True, color=C_WHITE, size=12)
c.alignment = align("center")
row_height(ws4, 1, 22)

raw_hdrs = ["Program","ปีพุทธศักราช","รอบ","รหัสตัวอย่าง","ค่ากำหนด",
            "Type/Genotype","ผลห้องปฏิบัติการ","คะแนน","ผลรอบ","TP/TN/FN","CT (QCMD)","Peer % (QCMD)"]
for ci, h in enumerate(raw_hdrs, start=1):
    apply_header(ws4, 2, ci, h, bg="234F6E", fontsize=9)
row_height(ws4, 2, 16)

raw_data = [
    ("QCMD","2567","C1","HPVPRES24C1-01","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","29.70","99.0%"),
    ("QCMD","2567","C1","HPVPRES24C1-02","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","26.44","99.5%"),
    ("QCMD","2567","C1","HPVPRES24C1-03","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","26.53","99.0%"),
    ("QCMD","2567","C1","HPVPRES24C1-04","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","23.51","99.0%"),
    ("QCMD","2567","C1","HPVPRES24C1-05","HPV16 Caski (EDU)","HPV16","HPV 16 (Positive)","0","PASS","TP","31.05","84.8%"),
    ("QCMD","2567","C1","HPVPRES24C1-06","Negative","Neg","Negative","0","PASS","TN","—","99.0%"),
    ("QCMD","2567","C2","HPVPRES24C2-01","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","25.49","98.5%"),
    ("QCMD","2567","C2","HPVPRES24C2-02","HPV45 (CC10b)","HPV45","HPV 45 (Positive)","0","PASS","TP","26.63","99.0%"),
    ("QCMD","2567","C2","HPVPRES24C2-03","Negative","Neg","Negative","0","PASS","TN","—","98.0%"),
    ("QCMD","2567","C2","HPVPRES24C2-04","HPV45 (CC10b)","HPV45","HPV 45 (Positive)","0","PASS","TP","26.54","98.5%"),
    ("QCMD","2567","C2","HPVPRES24C2-05","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","26.75","98.0%"),
    ("QCMD","2567","C2","HPVPRES24C2-06","HPV16+HPV18","HPV16+18","HPV 16+18 (Positive)","0","PASS","TP","27.16","99.0%"),
    ("QCMD","2568","C1","HPVPRES25C1-01","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","27.26","100.0%"),
    ("QCMD","2568","C1","HPVPRES25C1-02","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","27.37","98.8%"),
    ("QCMD","2568","C1","HPVPRES25C1-03","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","23.43","99.6%"),
    ("QCMD","2568","C1","HPVPRES25C1-04","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","27.21","99.6%"),
    ("QCMD","2568","C1","HPVPRES25C1-05","Negative","Neg","Negative","0","PASS","TN","—","98.0%"),
    ("QCMD","2568","C1","HPVPRES25C1-06","HPV18 low","HPV18","HPV 18 (Positive)","0","PASS","TP","30.15","94.9%"),
    ("QCMD","2568","C2","HPVPRES25C2-01","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","25.39","99.2%"),
    ("QCMD","2568","C2","HPVPRES25C2-02","HPV45 (CC10b)","HPV45","HPV 45 (Positive)","0","PASS","TP","27.70","98.4%"),
    ("QCMD","2568","C2","HPVPRES25C2-03","HPV18 (Hela)","HPV18","HPV 18 (Positive)","0","PASS","TP","26.48","98.0%"),
    ("QCMD","2568","C2","HPVPRES25C2-04","Negative","Neg","Negative","0","PASS","TN","—","98.8%"),
    ("QCMD","2568","C2","HPVPRES25C2-05","HPV45 (CC10b)","HPV45","HPV 45 (Positive)","0","PASS","TP","27.55","98.4%"),
    ("QCMD","2568","C2","HPVPRES25C2-06","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","24.91","99.2%"),
    ("QCMD","2569","C1","HPVPRES26C1-01","HPV16 Caski (low)","HPV16","HPV 16 (Positive)","0","PASS","TP","28.41","96.9%"),
    ("QCMD","2569","C1","HPVPRES26C1-02","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","25.63","99.1%"),
    ("QCMD","2569","C1","HPVPRES26C1-03","HPV16 (Caski)","HPV16","HPV 16 (Positive)","0","PASS","TP","26.16","99.1%"),
    ("QCMD","2569","C1","HPVPRES26C1-04","HPV45 (CC10b)","HPV45","HPV 45 (Positive)","0","PASS","TP","26.64","98.7%"),
    ("QCMD","2569","C1","HPVPRES26C1-05","HPV16 Caski (low)","HPV16","HPV 16 (Positive)","0","PASS","TP","29.99","97.3%"),
    ("QCMD","2569","C1","HPVPRES26C1-06","Negative","Neg","Negative","0","PASS","TN","—","98.2%"),
    ("สวส.","2567","C1","HPV-01","HPV 52 (Non-16,18)","HPV Non-16,18","HPV Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2567","C1","HPV-02","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2567","C1","HPV-03","HPV 16 + HPV 66","HPV16 + Non-16,18","HPV16 + Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2567","C1","HPV-04","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2567","C1","HPV-05","HPV 58 + HPV 39 (ยกเว้น)","HPV Non-16,18","HPV Non-16,18","—","PASS","EXCL","—","—"),
    ("สวส.","2567","C2","HPV-01","HPV 16","HPV16","HPV 16","2/2","PASS","TP","—","—"),
    ("สวส.","2567","C2","HPV-02","HPV 52 (Non-16,18)","HPV Non-16,18","HPV Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2567","C2","HPV-03","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2567","C2","HPV-04","HPV 68 (Non-16,18)","HPV Non-16,18","HPV Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2567","C2","HPV-05","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2568","C1","HPV-01","HPV 66 (Non-16,18)","HPV Non-16,18","Negative","0/2","FAIL","FN","—","—"),
    ("สวส.","2568","C1","HPV-02","Negative","Negative","Negative","2/2","FAIL","TN","—","—"),
    ("สวส.","2568","C1","HPV-03","HPV 16","HPV16","HPV 16","2/2","FAIL","TP","—","—"),
    ("สวส.","2568","C1","HPV-04","HPV 18","HPV18","HPV 18","2/2","FAIL","TP","—","—"),
    ("สวส.","2568","C1","HPV-05","HPV 18","HPV18","HPV 18","2/2","FAIL","TP","—","—"),
    ("สวส.","2568","C2","HPV-01","HPV 52 (Non-16,18)","HPV Non-16,18","HPV Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2568","C2","HPV-02","HPV 18","HPV18","HPV 18","2/2","PASS","TP","—","—"),
    ("สวส.","2568","C2","HPV-03","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2568","C2","HPV-04","HPV 58 (Non-16,18) ยกเว้น","HPV Non-16,18","Negative","—","PASS","EXCL","—","—"),
    ("สวส.","2568","C2","HPV-05","HPV 18","HPV18","HPV 18","2/2","PASS","TP","—","—"),
    ("สวส.","2569","C1","HPV-01","HPV 16","HPV16","HPV 16","2/2","PASS","TP","—","—"),
    ("สวส.","2569","C1","HPV-02","Negative","Negative","Negative","2/2","PASS","TN","—","—"),
    ("สวส.","2569","C1","HPV-03","HPV 18","HPV18","HPV 18","2/2","PASS","TP","—","—"),
    ("สวส.","2569","C1","HPV-04","HPV 52 (Non-16,18)","HPV Non-16,18","HPV Non-16,18","2/2","PASS","TP","—","—"),
    ("สวส.","2569","C1","HPV-05","HPV 18","HPV18","HPV 18","2/2","PASS","TP","—","—"),
]

for i, rd in enumerate(raw_data):
    r = 3 + i
    is_fn = rd[9] == "FN"
    is_excl = rd[9] == "EXCL"
    is_fail_round = rd[8] == "FAIL"
    is_svs = rd[0] == "สวส."
    row_bg2 = C_FAIL_BG if is_fn else (C_EXCL_BG if is_excl else (C_LGRAY if i % 2 == 0 else None))

    prog_bg = C_EXCL_BG if is_svs else "E0F7F4"
    prog_fg = C_PURPLE if is_svs else C_TEAL
    apply_cell(ws4, r, 1, rd[0], bg=prog_bg, fg=prog_fg, bold=True)

    for ci, val in enumerate(rd[1:], start=2):
        col_bg = row_bg2
        col_fg = "000000"
        col_bold = False

        if ci == 9:  # ผลรอบ
            col_bg = C_PASS_BG if val == "PASS" else C_FAIL_BG
            col_fg = C_PASS_FG if val == "PASS" else C_FAIL_FG
            col_bold = True
        elif ci == 10:  # TP/TN/FN
            verdict_colors = {"TP":(C_PASS_BG, C_PASS_FG), "TN":(C_HPV16,"154360"),
                              "FN":(C_FAIL_BG, C_FAIL_FG), "EXCL":(C_EXCL_BG, C_EXCL_FG)}
            col_bg, col_fg = verdict_colors.get(val, (row_bg2, "000000"))
            col_bold = True

        apply_cell(ws4, r, ci, val, bg=col_bg, fg=col_fg, bold=col_bold,
                   h="center" if ci in [2,3,8,9,10,11,12] else "left")
    row_height(ws4, r, 14)

set_col_widths(ws4, {1:10,2:8,3:6,4:22,5:28,6:16,7:28,8:8,9:10,10:10,11:10,12:12})

# ─── Save ────────────────────────────────────────────────────────────────────
output_path = "/sessions/intelligent-amazing-feynman/mnt/outputs/EQA_HPV_HB16305_Combined.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
