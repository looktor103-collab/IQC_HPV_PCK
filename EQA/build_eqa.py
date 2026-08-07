import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import copy

# ─── colour palette ───────────────────────────────────────────────
C_HEADER_DARK  = "1F4E79"   # dark blue
C_HEADER_MID   = "2E75B6"   # mid blue
C_HEADER_LIGHT = "D6E4F7"   # pale blue
C_QCMD_DARK    = "203864"   # navy for QCMD
C_QCMD_MID     = "375E97"
C_QCMD_LIGHT   = "D9E1F2"
C_GREEN_FILL   = "E2EFDA"
C_GREEN_DARK   = "375623"
C_WARN_FILL    = "FFF2CC"
C_WARN_DARK    = "7F6000"
C_RED_FILL     = "FFE2E2"
C_ROW_ALT      = "EBF3FB"
WHITE          = "FFFFFF"
C_SCORE_GOOD   = "C6EFCE"
C_SCORE_WARN   = "FFEB9C"
C_SCORE_BAD    = "FFC7CE"
C_SECTION      = "BDD7EE"

thin  = Side(style="thin",  color="BFBFBF")
thick = Side(style="medium", color="4472C4")
no    = Side(style=None)
THIN_BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_BORDER = Border(left=thick,right=thick,top=thick,bottom=thick)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Angsana New" if size>=13 else "Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, bold=False, size=10, color="000000",
             bg=None, halign="left", wrap=False, italic=False, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=size, color=color, italic=italic,
                       name="Calibri")
    c.alignment = Alignment(horizontal=halign, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if border:
        c.border = border
    return c

def merge_set(ws, r1, c1, r2, c2, value, bold=False, size=10,
              color="000000", bg=None, halign="center", wrap=False,
              italic=False):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font      = Font(bold=bold, size=size, color=color, italic=italic,
                       name="Calibri")
    c.alignment = Alignment(horizontal=halign, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    return c

def score_fill(score_int):
    if score_int == 0:  return C_SCORE_GOOD
    if score_int == 1:  return C_SCORE_GOOD
    if score_int == 2:  return C_SCORE_WARN
    return C_SCORE_BAD

def qcmd_label(score_int):
    if score_int == 0: return "✅ ดีเยี่ยม (≤1 SD)"
    if score_int == 1: return "✅ ยอมรับ (1–2 SD)"
    if score_int == 2: return "⚠️ เฝ้าระวัง (2–3 SD)"
    return "❌ ไม่ผ่าน (>3 SD)"

# ═══════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)     # remove default sheet

# ───────────────────────────────────────────────────────────────────
# helpers to build standard detail sheet (สวส.)
# ───────────────────────────────────────────────────────────────────
def build_svs_sheet(wb, sheet_name, title, date_str, report_no, lab_code,
                    assay, samples, precision_label, blab_z, wlab_z,
                    total_score, note=None):
    """
    samples = [(code, xpt_log, result_log, copies_or_dash, zscore, label), ...]
    precision_label = e.g. "HIVL 682-1"
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    r = 1
    # Title row
    merge_set(ws, r, 1, r, 8, title, bold=True, size=14,
              color=WHITE, bg=C_HEADER_DARK, halign="center")
    ws.row_dimensions[r].height = 26
    r += 1

    meta = [("วันที่ออกรายงาน:", date_str),
            ("เลขรายงาน:", report_no),
            ("รหัสสมาชิก:", lab_code),
            ("ชุดตรวจ:", assay)]
    if note:
        meta.append(("หมายเหตุ:", note))

    for label, val in meta:
        set_cell(ws, r, 1, label, bold=True, bg=C_HEADER_LIGHT)
        merge_set(ws, r, 2, r, 4, val)
        r += 1

    r += 1  # blank

    # Accuracy section
    merge_set(ws, r, 1, r, 8, "📊 ผลความแม่น (Accuracy)",
              bold=True, size=11, color=WHITE, bg=C_HEADER_MID,
              halign="left")
    r += 1

    hdrs = ["รหัสตัวอย่าง",
            "ค่ากำหนด Xpt\n(Log₁₀)",
            "ผลทดสอบ\n(Log₁₀)",
            "copies/mL",
            "Z-score",
            "ผลการแปล"]
    for ci, h in enumerate(hdrs, 1):
        set_cell(ws, r, ci, h, bold=True, bg=C_SECTION, halign="center",
                 wrap=True, border=THIN_BORDER)
    ws.row_dimensions[r].height = 36
    r += 1

    for i, (code, xpt, res, cpm, z, lbl) in enumerate(samples):
        bg = WHITE if i % 2 == 0 else C_ROW_ALT
        set_cell(ws, r, 1, code, bg=bg, border=THIN_BORDER)
        set_cell(ws, r, 2, xpt,  bg=bg, halign="center", border=THIN_BORDER)
        set_cell(ws, r, 3, res,  bg=bg, halign="center", border=THIN_BORDER)
        set_cell(ws, r, 4, cpm,  bg=bg, halign="center", border=THIN_BORDER)
        set_cell(ws, r, 5, z,    bg=bg, halign="center", border=THIN_BORDER)
        # colour label
        if "ยอมรับ" in str(lbl):    lbl_bg = C_SCORE_GOOD
        elif "เฝ้าระวัง" in str(lbl): lbl_bg = C_SCORE_WARN
        elif "ออกนอก" in str(lbl):   lbl_bg = C_SCORE_BAD
        else: lbl_bg = bg
        set_cell(ws, r, 6, lbl, bg=lbl_bg, halign="center",
                 border=THIN_BORDER)
        r += 1

    r += 1
    # Precision section
    prec_title = f"📐 ผลความเที่ยง (Precision) — {precision_label}"
    merge_set(ws, r, 1, r, 8, prec_title,
              bold=True, size=11, color=WHITE, bg=C_HEADER_MID,
              halign="left")
    r += 1

    def prec_row(label, z_val, interp):
        if "ยอมรับ" in str(interp):    ibg = C_SCORE_GOOD
        elif "เฝ้าระวัง" in str(interp): ibg = C_SCORE_WARN
        elif "ออกนอก" in str(interp):   ibg = C_SCORE_BAD
        else: ibg = WHITE
        merge_set(ws, r, 1, r, 2, label, bold=True, bg=C_HEADER_LIGHT)
        set_cell(ws, r, 3, z_val,  halign="center", border=THIN_BORDER)
        set_cell(ws, r, 4, interp, bg=ibg, halign="center", border=THIN_BORDER)

    prec_row("Between-lab z-score", blab_z,
             "✅ ยอมรับ" if blab_z not in (None,"—","NA","N/A") and abs(float(str(blab_z).replace("+",""))) < 2 else ("⚠️ เฝ้าระวัง" if blab_z not in (None,"—","NA","N/A") and abs(float(str(blab_z).replace("+",""))) < 3 else "❌ ออกนอกเกณฑ์") if blab_z not in (None,"—","NA","N/A") else "—")
    r += 1
    prec_row("Within-lab z-score",  wlab_z,
             "✅ ยอมรับ" if wlab_z not in (None,"—","NA","N/A") and abs(float(str(wlab_z).replace("+",""))) < 2 else ("⚠️ เฝ้าระวัง" if wlab_z not in (None,"—","NA","N/A") and abs(float(str(wlab_z).replace("+",""))) < 3 else "❌ ออกนอกเกณฑ์") if wlab_z not in (None,"—","NA","N/A") else "—")
    r += 1

    r += 1
    merge_set(ws, r, 1, r, 2, "คะแนนรวม", bold=True, bg=C_HEADER_LIGHT)
    ts_bg = C_SCORE_GOOD if total_score == "20/20" else (C_SCORE_WARN if total_score == "19/20" else WHITE)
    merge_set(ws, r, 3, r, 4, total_score, bold=True, bg=ts_bg,
              size=12)
    return ws

# ───────────────────────────────────────────────────────────────────
# build QCMD detail sheet
# ───────────────────────────────────────────────────────────────────
def build_qcmd_sheet(wb, sheet_name, title, issue_date, ref_code, challenge,
                     lab_id, assay_group, qual_score, quant_score,
                     panel, note=None):
    """
    panel = [(code, status, consensus_log, sd, your_log, est_score,
              is_negative), ...]
    is_negative=True -> negative control, no quant
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 14

    r = 1
    merge_set(ws, r, 1, r, 8, title, bold=True, size=14,
              color=WHITE, bg=C_QCMD_DARK, halign="center")
    ws.row_dimensions[r].height = 26
    r += 1

    meta = [("วันที่ออกรายงาน:", issue_date),
            ("Ref Code / Challenge:", f"{ref_code} / {challenge}"),
            ("Laboratory ID:", lab_id),
            ("EQA Assessment Group:", assay_group)]
    if note:
        meta.append(("หมายเหตุ:", note))

    for label, val in meta:
        set_cell(ws, r, 1, label, bold=True, bg=C_QCMD_LIGHT)
        merge_set(ws, r, 2, r, 5, val)
        r += 1

    r += 1
    # Summary scores box
    merge_set(ws, r, 1, r, 8, "📊 ผลคะแนนรวม (Summary Scores)",
              bold=True, size=11, color=WHITE, bg=C_QCMD_MID, halign="left")
    r += 1
    set_cell(ws, r, 1, "Core Detection Score (Qualitative):", bold=True,
             bg=C_QCMD_LIGHT)
    qs_val = qual_score if qual_score is not None else "N/A"
    set_cell(ws, r, 2, qs_val, halign="center", bg=WHITE,
             border=THIN_BORDER)
    set_cell(ws, r, 4, "Core Estimation Score (Quantitative):", bold=True,
             bg=C_QCMD_LIGHT)
    qt_bg = score_fill(quant_score) if isinstance(quant_score, int) else WHITE
    set_cell(ws, r, 5, str(quant_score) if quant_score is not None else "N/A",
             halign="center", bg=qt_bg, bold=True, border=THIN_BORDER)
    # QCMD score interpretation note
    if isinstance(quant_score, int):
        set_cell(ws, r, 6, qcmd_label(quant_score), bg=qt_bg,
                 border=THIN_BORDER)
    r += 1

    r += 1
    # Panel results table
    merge_set(ws, r, 1, r, 8, "📋 ผลการทดสอบตัวอย่าง (Panel Results)",
              bold=True, size=11, color=WHITE, bg=C_QCMD_MID, halign="left")
    r += 1

    hdrs = ["Sample Code", "Status", "Consensus\n(Log₁₀)",
            "SD", "Your Result\n(Log₁₀)", "Est. Score", "การแปลผล",
            "หมายเหตุ"]
    for ci, h in enumerate(hdrs, 1):
        set_cell(ws, r, ci, h, bold=True, bg=C_SECTION,
                 halign="center", wrap=True, border=THIN_BORDER)
    ws.row_dimensions[r].height = 32
    r += 1

    for i, row_data in enumerate(panel):
        code, status, cons, sd, your_res, est, is_neg = row_data
        bg = WHITE if i % 2 == 0 else C_QCMD_LIGHT

        set_cell(ws, r, 1, code,    bg=bg, border=THIN_BORDER)
        status_bg = C_SCORE_WARN if status=="EDUCATIONAL" else bg
        set_cell(ws, r, 2, status,  bg=status_bg, halign="center",
                 border=THIN_BORDER, italic=(status=="EDUCATIONAL"))

        if is_neg:
            set_cell(ws, r, 3, "N/A (Negative)", bg=bg, halign="center",
                     border=THIN_BORDER, italic=True)
            set_cell(ws, r, 4, "—",   bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 5, "LOD/NR", bg=bg, halign="center",
                     border=THIN_BORDER, italic=True)
            set_cell(ws, r, 6, "N/A", bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 7, "ตรวจไม่พบ (Expected)", bg=C_SCORE_GOOD,
                     halign="center", border=THIN_BORDER)
        elif status == "EDUCATIONAL":
            set_cell(ws, r, 3, f"{cons:.3f}" if cons else "—", bg=bg,
                     halign="center", border=THIN_BORDER)
            set_cell(ws, r, 4, f"{sd:.3f}" if sd else "—", bg=bg,
                     halign="center", border=THIN_BORDER)
            set_cell(ws, r, 5, f"{your_res:.3f}" if your_res else "—",
                     bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 6, "—",  bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 7, "—",  bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 8, "Educational (ไม่คิดคะแนน)",
                     bg=C_SCORE_WARN, halign="center", border=THIN_BORDER,
                     italic=True)
        else:
            set_cell(ws, r, 3, f"{cons:.3f}" if cons else "—", bg=bg,
                     halign="center", border=THIN_BORDER)
            set_cell(ws, r, 4, f"{sd:.3f}" if sd else "—", bg=bg,
                     halign="center", border=THIN_BORDER)
            set_cell(ws, r, 5, f"{your_res:.3f}" if your_res else "—",
                     bg=bg, halign="center", border=THIN_BORDER)
            est_bg = score_fill(est) if isinstance(est, int) else bg
            set_cell(ws, r, 6, str(est) if isinstance(est, int) else "—",
                     bg=est_bg, halign="center", bold=True, border=THIN_BORDER)
            set_cell(ws, r, 7, qcmd_label(est) if isinstance(est, int) else "—",
                     bg=est_bg, halign="center", border=THIN_BORDER)
        r += 1

    r += 1
    merge_set(ws, r, 1, r, 4,
              "⚠️ QCMD Estimation Score: 0=ดีเยี่ยม(≤1SD), 1=ยอมรับ(1-2SD), 2=เฝ้าระวัง(2-3SD), 3=ไม่ผ่าน(>3SD)",
              italic=True, size=9, color="595959", halign="left")
    return ws


# ───────────────────────────────────────────────────────────────────
# build QCMD online-only sheet (2023 — no official individual report)
# ───────────────────────────────────────────────────────────────────
def build_qcmd_online_sheet(wb, sheet_name, title, submit_date, programme,
                             challenge, samples_cpm):
    """samples_cpm = [(code, copies_ml or 'ND')]"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    r = 1
    merge_set(ws, r, 1, r, 4, title, bold=True, size=14,
              color=WHITE, bg=C_QCMD_DARK, halign="center")
    ws.row_dimensions[r].height = 26
    r += 1

    meta = [("วันที่ส่งผล:", submit_date),
            ("Programme:", programme),
            ("Challenge:", challenge),
            ("หมายเหตุ:", "รายงานเบื้องต้นจาก QCMD ITEMS (online) — ไม่มีรายงานรายห้องอย่างเป็นทางการ")]
    for label, val in meta:
        set_cell(ws, r, 1, label, bold=True, bg=C_QCMD_LIGHT)
        merge_set(ws, r, 2, r, 4, val, italic=(label=="หมายเหตุ:"),
                  color="595959" if label=="หมายเหตุ:" else "000000")
        r += 1

    r += 1
    merge_set(ws, r, 1, r, 4, "📋 ผลที่ส่ง (Submitted Results)",
              bold=True, size=11, color=WHITE, bg=C_QCMD_MID, halign="left")
    r += 1

    for h, ci in [("Sample Code",1),("Copies/mL",2),("Log₁₀",3),("หมายเหตุ",4)]:
        set_cell(ws, r, ci, h, bold=True, bg=C_SECTION,
                 halign="center", border=THIN_BORDER)
    r += 1

    for i, (code, cpm) in enumerate(samples_cpm):
        bg = WHITE if i%2==0 else C_QCMD_LIGHT
        set_cell(ws, r, 1, code, bg=bg, border=THIN_BORDER)
        if cpm in (0,"0","ND","Target Not Detected","not detected"):
            set_cell(ws, r, 2, "Not Detected", bg=bg, halign="center",
                     border=THIN_BORDER, italic=True)
            set_cell(ws, r, 3, "—", bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 4, "ND", bg=bg, halign="center", border=THIN_BORDER)
        else:
            import math
            cpm_n = float(cpm)
            log_v = round(math.log10(cpm_n), 3) if cpm_n > 0 else "—"
            set_cell(ws, r, 2, cpm_n, bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 3, log_v, bg=bg, halign="center", border=THIN_BORDER)
            set_cell(ws, r, 4, "", bg=bg, border=THIN_BORDER)
        r += 1

    r += 1
    merge_set(ws, r, 1, r, 4,
              "⚠️ ไม่มีรายงานผลรายห้องอย่างเป็นทางการ (Individual Report) จาก QCMD สำหรับรอบนี้",
              italic=True, size=9, color="7F0000", halign="left")
    return ws


# ═══════════════════════════════════════════════════════════════════
#  BUILD ALL DETAIL SHEETS
# ═══════════════════════════════════════════════════════════════════

# ── 2569 / 2026 ────────────────────────────────────────────────────
build_svs_sheet(wb,
    "📋 สวส. C1-2569",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 1-2569 (สวส.)",
    "20 มีนาคม 2569", "69-047", "00031", "Cobas HIV-1",
    [("HIVL 691-1", "3.46 ± 0.17", "2.88",  "—", "+0.00", "✅ ยอมรับ"),
     ("HIVL 691-2", "4.96 ± 0.17", "4.91",  "81,100", "-0.29", "✅ ยอมรับ"),
     ("HIVL 691-3", "4.94 ± 0.17", "4.78",  "60,400", "-0.94", "✅ ยอมรับ"),
     ("HIVL 691-4", "5.54 ± 0.17", "5.36",  "228,000", "-1.06", "✅ ยอมรับ")],
    "HIVL 691-2 vs 691-3", "-0.88", "+1.11", "20/20"
)

build_qcmd_sheet(wb,
    "📋 QCMD 2026 C1",
    "ผลการประเมิน QCMD 2026 HIV RNA EQA — Challenge C1",
    "19 May 2026", "HIVRNA26", "C1", "TH193",
    "Roche Cobas 5800/6800/8800",
    qual_score=None, quant_score=2,
    panel=[
        # code, status, consensus, sd, your_res, est_score, is_neg
        ("HIVRNA26C1-01", "CORE",        2.865, 0.155, 2.708, 1, False),
        ("HIVRNA26C1-02", "CORE",        3.877, 0.156, 3.659, 1, False),
        ("HIVRNA26C1-03", "EDUCATIONAL", 1.865, None,  None,  None, False),
        ("HIVRNA26C1-04", "CORE",        None,  None,  None,  None, True),
    ],
    note="Sample 03 เป็น Educational ไม่คิดคะแนน"
)

# ── 2568 / 2025 ────────────────────────────────────────────────────
build_svs_sheet(wb,
    "📋 สวส. C1-2568",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 1-2568 (สวส.)",
    "21 มีนาคม 2568", "68-025", "00044", "Cobas HIV-1",
    [("HIVL 681-1", "3.46 ± 0.17", "3.21", "—", "-1.47", "✅ ยอมรับ"),
     ("HIVL 681-2", "—",           "ND",   "—", "—",     "ND"),
     ("HIVL 681-3", "3.96 ± 0.17", "3.82", "—", "-0.82", "✅ ยอมรับ"),
     ("HIVL 681-4", "3.99 ± 0.17", "3.99", "—", "+0.00", "✅ ยอมรับ")],
    "HIVL 681-3", "-0.50", "-1.38", "20/20"
)

build_svs_sheet(wb,
    "📋 สวส. C2-2568",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 2-2568 (สวส.)",
    "20 สิงหาคม 2568", "68-074", "00044", "Cobas HIV-1",
    [("HIVL 682-1", "3.96 ± 0.17", "3.86", "—", "-0.59", "✅ ยอมรับ"),
     ("HIVL 682-2", "3.93 ± 0.17", "3.97", "—", "+0.24", "✅ ยอมรับ"),
     ("HIVL 682-3", "3.49 ± 0.17", "3.48", "—", "-0.06", "✅ ยอมรับ"),
     ("HIVL 682-4", "5.24 ± 0.17", "5.20", "—", "-0.24", "✅ ยอมรับ")],
    "HIVL 682-1", "-0.14", "-2.50", "19/20",
    note="Within-lab z = −2.50 อยู่ในเกณฑ์เฝ้าระวัง"
)

build_qcmd_sheet(wb,
    "📋 QCMD 2025 C1",
    "ผลการประเมิน QCMD 2025 HIV RNA EQA — Challenge C1",
    "03 April 2025", "HIVRNA25", "C1", "TH193",
    "Roche Cobas 5800/6800/8800",
    qual_score=None, quant_score=0,
    panel=[
        ("HIVRNA25C1-01", "CORE",        2.729, 0.169, 2.696, 0, False),
        ("HIVRNA25C1-02", "CORE",        3.768, 0.165, 3.626, 0, False),
        ("HIVRNA25C1-03", "EDUCATIONAL", 1.717, None,  None,  None, False),
        ("HIVRNA25C1-04", "CORE",        2.722, 0.172, 2.651, 0, False),
    ]
)

build_qcmd_sheet(wb,
    "📋 QCMD 2025 C3",
    "ผลการประเมิน QCMD 2025 HIV RNA EQA — Challenge C3",
    "21 August 2025", "HIVRNA25", "C3", "TH193",
    "Roche Cobas 5800/6800/8800",
    qual_score=None, quant_score=1,
    panel=[
        ("HIVRNA25C3-01", "CORE", 2.823, 0.137, 2.862, 0, False),
        ("HIVRNA25C3-02", "CORE", None,  None,  None,  None, True),
        ("HIVRNA25C3-03", "CORE", 2.807, 0.173, 2.573, 1, False),
        ("HIVRNA25C3-04", "CORE", 2.803, 0.143, 2.828, 0, False),
    ]
)

# Certificate sheet
ws_cert = wb.create_sheet("🏆 Certificate QCMD 2025")
ws_cert.sheet_view.showGridLines = False
ws_cert.column_dimensions["A"].width = 100
merge_set(ws_cert, 1, 1, 1, 6,
          "ใบรับรองการเข้าร่วม QCMD 2025 HIV RNA EQA Scheme — Laboratory TH193",
          bold=True, size=13, color=WHITE, bg=C_QCMD_DARK, halign="center")
ws_cert.row_dimensions[1].height = 26
set_cell(ws_cert, 2, 1,
         "QCMD 2025 Human Immunodeficiency Virus RNA EQA Scheme — QAV994108_2",
         bold=True, size=11)
set_cell(ws_cert, 3, 1,
         "Laboratory: Phrachomklao Hospital, Phetchaburi, Thailand")
set_cell(ws_cert, 4, 1,
         "Certificate valid: 21st August 2025 to 21st August 2026",
         bold=True, color="1F4E79")

img = XLImage("/sessions/practical-sweet-curie/mnt/outputs/cert2025.png")
img.width  = 600
img.height = 849
img.anchor = "A6"
ws_cert.add_image(img)

# ── 2567 / 2024 ────────────────────────────────────────────────────
build_svs_sheet(wb,
    "📋 สวส. C1-2567",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 1-2567 (สวส.)",
    "28 มีนาคม 2567", "67-021", "00039", "Cobas HIV-1",
    [("HIVL 671-1", "4.58 ± 0.17", "4.44",    "—", "-0.82", "✅ ยอมรับ"),
     ("HIVL 671-2", "2.71 ± 0.17", "2.70",    "—", "-0.06", "✅ ยอมรับ"),
     ("HIVL 671-3", "2.70 ± 0.17", "2.49",    "—", "-1.24", "✅ ยอมรับ"),
     ("HIVL 671-4", "—",           "Negative","—", "—",     "Negative")],
    "HIVL 671-2 vs 671-3", "-1.33", "+0.50", "20/20"
)

build_svs_sheet(wb,
    "📋 สวส. C2-2567",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 2-2567 (สวส.)",
    "16 สิงหาคม 2567", "67-074", "00039", "Cobas HIV-1",
    [("HIVL 672-1", "3.75 ± 0.17", "3.63", "—", "-0.71", "✅ ยอมรับ"),
     ("HIVL 672-2", "4.62 ± 0.17", "4.50", "—", "-0.71", "✅ ยอมรับ"),
     ("HIVL 672-3", "4.63 ± 0.17", "4.47", "—", "-0.94", "✅ ยอมรับ"),
     ("HIVL 672-4", "5.76 ± 0.17", "5.51", "—", "-1.47", "✅ ยอมรับ")],
    "HIVL 672-2 vs 672-3", "-1.00", "+0.29", "20/20"
)

# ── 2566 / 2023 ────────────────────────────────────────────────────
build_svs_sheet(wb,
    "📋 สวส. C1-2566",
    "ผลการประเมิน EQA HIV-1 Viral Load — ครั้งที่ 1-2566 (สวส.)",
    "10 มีนาคม 2566", "66-001", "00044", "Cobas HIV-1",
    [("HIVL 661-1", "3.82 ± 0.17", "3.80", "—", "-0.41", "✅ ยอมรับ"),
     ("HIVL 661-2", "4.33 ± 0.17", "4.23", "—", "-0.87", "✅ ยอมรับ"),
     ("HIVL 661-3", "3.81 ± 0.17", "3.88", "—", "+0.69", "✅ ยอมรับ"),
     ("HIVL 661-4", "5.34 ± 0.17", "5.24", "—", "-1.27", "✅ ยอมรับ"),
     ("HIVL 661-5", "—",           "ND",   "—", "—",     "ND"),
     ("HIVL 661-6", "2.80 ± 0.17", "2.72", "—", "-0.75", "✅ ยอมรับ")],
    "—", "—", "—", "—",
    note="⚠️ รายงานเบื้องต้น — ไม่มีข้อมูล Precision"
)

build_qcmd_online_sheet(wb,
    "📋 QCMD 2023 C1",
    "ผลการส่ง QCMD 2023 HIV RNA EQA — Challenge C1 (Online Submission)",
    "10 มีนาคม 2566", "QCMD 2023 HIV RNA EQA (HIVRNA23)",
    "C1",
    [("HIVRNA23C1-01", 1570),
     ("HIVRNA23C1-02", "ND"),
     ("HIVRNA23C1-03", 958),
     ("HIVRNA23C1-04", 348)]
)

build_qcmd_online_sheet(wb,
    "📋 QCMD 2023 C3",
    "ผลการส่ง QCMD 2023 HIV RNA EQA — Challenge C3 (Online Submission)",
    "25 กรกฎาคม 2566", "QCMD 2023 HIV RNA EQA (HIVRNA23)",
    "C3",
    [("HIVRNA23C3-01", 49),
     ("HIVRNA23C3-02", 3270),
     ("HIVRNA23C3-03", 557),
     ("HIVRNA23C3-04", 644)]
)

# ═══════════════════════════════════════════════════════════════════
#  SUMMARY SHEET (first)
# ═══════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("📊 สรุปภาพรวม EQA", 0)
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 22
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 12
ws_sum.column_dimensions["E"].width = 12
ws_sum.column_dimensions["F"].width = 12
ws_sum.column_dimensions["G"].width = 12
ws_sum.column_dimensions["H"].width = 14
ws_sum.column_dimensions["I"].width = 14
ws_sum.column_dimensions["J"].width = 18

r = 1
merge_set(ws_sum, r, 1, r, 10,
    "สรุปผลการประเมิน EQA — HIV-1 Viral Load · โรงพยาบาลพระจอมเกล้า เพชรบุรี",
    bold=True, size=14, color=WHITE, bg=C_HEADER_DARK, halign="center")
ws_sum.row_dimensions[r].height = 28
r += 1

# ─── สวส. sub-table ───────────────────────────────────────────────
r += 1
merge_set(ws_sum, r, 1, r, 10,
    "🏛️  สถาบันวิจัยวิทยาศาสตร์สาธารณสุข (สวส.) กรมวิทยาศาสตร์การแพทย์",
    bold=True, size=12, color=WHITE, bg=C_HEADER_MID, halign="left")
r += 1
merge_set(ws_sum, r, 1, r, 10,
    "เกณฑ์: ISO 13528:2022 Algorithm A | σpt = 0.17 | |z| ≤ 2.00 = ยอมรับ | 2.00 < |z| < 3.00 = เฝ้าระวัง | |z| ≥ 3.00 = ออกนอกเกณฑ์",
    size=9, italic=True, color="595959")
r += 1

svs_hdrs = ["รอบการประเมิน","วันที่ออกรายงาน","รหัสสมาชิก",
            "z-1","z-2","z-3","z-4","Between-lab z","Within-lab z","คะแนน"]
for ci, h in enumerate(svs_hdrs, 1):
    set_cell(ws_sum, r, ci, h, bold=True, bg=C_SECTION,
             halign="center", wrap=True, border=THIN_BORDER)
ws_sum.row_dimensions[r].height = 30
r += 1

svs_rows = [
    ("ครั้งที่ 1-2569","20 มีนาคม 2569","00031","+0.00","-0.29","-0.94","-1.06","-0.88","+1.11","20/20"),
    ("ครั้งที่ 1-2568","21 มีนาคม 2568","00044","-1.47","—","-0.82","+0.00","-0.50","-1.38","20/20"),
    ("ครั้งที่ 2-2568","20 สิงหาคม 2568","00044","-0.59","+0.24","-0.06","-0.24","-0.14","-2.50⚠️","19/20"),
    ("ครั้งที่ 1-2567","28 มีนาคม 2567","00039","-0.82","-0.06","-1.24","—","-1.33","+0.50","20/20"),
    ("ครั้งที่ 2-2567","16 สิงหาคม 2567","00039","-0.71","-0.71","-0.94","-1.47","-1.00","+0.29","20/20"),
    ("ครั้งที่ 1-2566","10 มีนาคม 2566","00044","-0.41","-0.87","+0.69","-1.27","—","—","—"),
]
for i, row in enumerate(svs_rows):
    bg = WHITE if i%2==0 else C_ROW_ALT
    for ci, v in enumerate(row, 1):
        cbg = bg
        if ci==10:
            cbg = C_SCORE_GOOD if v=="20/20" else (C_SCORE_WARN if v=="19/20" else bg)
        set_cell(ws_sum, r, ci, v, bg=cbg, halign="center", border=THIN_BORDER,
                 bold=(ci==10))
    r += 1

# ─── QCMD sub-table ───────────────────────────────────────────────
r += 2
merge_set(ws_sum, r, 1, r, 10,
    "🌐  QCMD (Quality Control for Molecular Diagnostics) — Glasgow, UK",
    bold=True, size=12, color=WHITE, bg=C_QCMD_DARK, halign="left")
r += 1
merge_set(ws_sum, r, 1, r, 10,
    "เกณฑ์: Estimation Score — 0 = ดีเยี่ยม (≤1 SD) | 1 = ยอมรับ (1–2 SD) | 2 = เฝ้าระวัง (2–3 SD) | 3 = ไม่ผ่าน (>3 SD)  |  Laboratory ID: TH193",
    size=9, italic=True, color="595959")
r += 1

qcmd_hdrs = ["รอบการประเมิน","วันที่ออกรายงาน","EQA Group",
             "Quantitative Score","สรุปผล","","","","",""]
for ci, h in enumerate(qcmd_hdrs[:5], 1):
    set_cell(ws_sum, r, ci, h, bold=True, bg=C_SECTION,
             halign="center", wrap=True, border=THIN_BORDER)
for ci in range(6,11):
    ws_sum.cell(row=r, column=ci).fill = fill(C_SECTION)
merge_set(ws_sum, r, 5, r, 10, "การแปลผล", bold=True, bg=C_SECTION,
          halign="center")
ws_sum.row_dimensions[r].height = 30
r += 1

qcmd_rows = [
    ("HIVRNA26 C1 (2026)","19 May 2026","Roche Cobas\n5800/6800/8800", 2, "✅ ยอมรับ"),
    ("HIVRNA25 C1 (2025)","03 Apr 2025","Roche Cobas\n5800/6800/8800", 0, "✅ ดีเยี่ยม"),
    ("HIVRNA25 C3 (2025)","21 Aug 2025","Roche Cobas\n5800/6800/8800", 1, "✅ ยอมรับ"),
    ("HIVRNA23 C1 (2023)","10 Mar 2023 (submitted)","Cobas AmpliPrep/TaqMan", "—", "ไม่มีรายงานรายห้อง"),
    ("HIVRNA23 C3 (2023)","25 Jul 2023 (submitted)","Cobas 6800",            "—", "ไม่มีรายงานรายห้อง"),
]
for i, (rnd, dt, grp, sc, lbl) in enumerate(qcmd_rows):
    bg = WHITE if i%2==0 else C_QCMD_LIGHT
    set_cell(ws_sum, r, 1, rnd, bg=bg, border=THIN_BORDER)
    set_cell(ws_sum, r, 2, dt, bg=bg, halign="center", border=THIN_BORDER)
    set_cell(ws_sum, r, 3, grp, bg=bg, halign="center", border=THIN_BORDER, wrap=True)
    sc_bg = score_fill(sc) if isinstance(sc, int) else bg
    set_cell(ws_sum, r, 4, str(sc), bg=sc_bg, halign="center",
             bold=True, border=THIN_BORDER)
    lbl_bg = C_SCORE_GOOD if "✅" in lbl else (C_SCORE_WARN if "⚠️" in lbl else bg)
    merge_set(ws_sum, r, 5, r, 10, lbl, bg=lbl_bg, halign="center")
    r += 1

# Certificate note
r += 1
merge_set(ws_sum, r, 1, r, 10,
    "🏆 QCMD 2025 Participation Certificate — Valid: 21 Aug 2025 → 21 Aug 2026  |  Laboratory: Phrachomklao Hospital, Phetchaburi  |  ดูใบรับรองที่ Sheet '🏆 Certificate QCMD 2025'",
    bold=True, size=10, color=WHITE, bg=C_QCMD_MID, halign="left")
ws_sum.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════
#  CRITERIA SHEET (last)
# ═══════════════════════════════════════════════════════════════════
ws_crit = wb.create_sheet("📖 เกณฑ์การแปลผล")
ws_crit.sheet_view.showGridLines = False
for col, w in [("A",30),("B",22),("C",28),("D",22),("E",28)]:
    ws_crit.column_dimensions[col].width = w

r = 1
merge_set(ws_crit, r, 1, r, 5,
          "📖 เกณฑ์การแปลผล EQA HIV-1 Viral Load",
          bold=True, size=13, color=WHITE, bg=C_HEADER_DARK, halign="center")
ws_crit.row_dimensions[r].height = 26
r += 2

merge_set(ws_crit, r, 1, r, 5,
          "สวส. — Z-score Criteria (ISO 13528:2022 Algorithm A)",
          bold=True, size=11, color=WHITE, bg=C_HEADER_MID)
r += 1
hdrs = ["ระดับ","ช่วง |z|","ความหมาย","คะแนนความแม่น","การดำเนินการ"]
for ci, h in enumerate(hdrs, 1):
    set_cell(ws_crit, r, ci, h, bold=True, bg=C_SECTION,
             halign="center", border=THIN_BORDER)
r += 1
svs_crit = [
    ("✅ ยอมรับ","|z| ≤ 2.00","Acceptable","3 คะแนน/ตัวอย่าง","ปกติ",C_SCORE_GOOD),
    ("⚠️ เฝ้าระวัง","2.00 < |z| < 3.00","Warning signal","2 คะแนน/ตัวอย่าง","ตรวจสอบและติดตาม",C_SCORE_WARN),
    ("❌ ออกนอกเกณฑ์","|z| ≥ 3.00","Action signal","1 คะแนน/ตัวอย่าง","วิเคราะห์หาสาเหตุและแก้ไข",C_SCORE_BAD),
]
for lvl, rng, meaning, pts, action, bg in svs_crit:
    for ci, v in enumerate([lvl,rng,meaning,pts,action], 1):
        set_cell(ws_crit, r, ci, v, bg=bg, halign="center", border=THIN_BORDER)
    r += 1

set_cell(ws_crit, r, 1,
    "คะแนนเต็ม 20 = รายงานผลทันเวลา (2) + ความแม่น 4×3 (12) + ความเที่ยง 2×3 (6)",
    italic=True, size=9, color="595959")
r += 2

merge_set(ws_crit, r, 1, r, 5,
          "QCMD — Estimation Score Criteria",
          bold=True, size=11, color=WHITE, bg=C_QCMD_DARK)
r += 1
hdrs2 = ["Score","ช่วง SD","ความหมาย","การแปลผล","หมายเหตุ"]
for ci, h in enumerate(hdrs2, 1):
    set_cell(ws_crit, r, ci, h, bold=True, bg=C_SECTION,
             halign="center", border=THIN_BORDER)
r += 1
qcmd_crit = [
    ("0","≤ 1 SD","Highly Satisfactory","✅ ดีเยี่ยม",C_SCORE_GOOD),
    ("1","1 – 2 SD","Satisfactory","✅ ยอมรับ",C_SCORE_GOOD),
    ("2","2 – 3 SD","Borderline","⚠️ เฝ้าระวัง",C_SCORE_WARN),
    ("3","> 3 SD","Unsatisfactory","❌ ไม่ผ่าน",C_SCORE_BAD),
]
for sc, rng, meaning, lbl, bg in qcmd_crit:
    for ci, v in enumerate([sc, rng, meaning, lbl, ""], 1):
        set_cell(ws_crit, r, ci, v, bg=bg, halign="center", border=THIN_BORDER)
    r += 1

set_cell(ws_crit, r, 1,
    "QCMD ไม่มีคะแนน Qualitative (Detection) ในโปรแกรมนี้ (N/A) — ประเมินเฉพาะ Quantitative Estimation",
    italic=True, size=9, color="595959")

# ═══════════════════════════════════════════════════════════════════
#  SAVE
# ═══════════════════════════════════════════════════════════════════
out = "/sessions/practical-sweet-curie/mnt/outputs/EQA_HIV_VL_All_Years_Updated.xlsx"
wb.save(out)
print("Saved:", out)
print("Sheets:", [ws.title for ws in wb.worksheets])
